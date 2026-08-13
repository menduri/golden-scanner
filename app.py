from flask import Flask, render_template, request, jsonify, send_file, Response
import yfinance as yf
import pandas as pd
import json, os, io, threading
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__)

# ── WATCHLIST SYNC (GitHub Gist) ─────────────────────────────────────────────
# Render's free tier wipes local disk on every restart/redeploy, so the
# watchlist can't live in a local file — it lives in a private GitHub Gist
# instead, which any device can read/write through the GitHub API.
# Fill these in after creating a token (Settings -> Developer settings ->
# Personal access tokens, "gist" scope) and a private Gist containing a file
# named "golden_scanner_watchlist.json" with content: []
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN', '')
GITHUB_GIST_ID = os.environ.get('GITHUB_GIST_ID', '')
GIST_FILENAME = 'golden_scanner_watchlist.json'
ANCHOR_FILENAME = 'golden_scanner_anchors.json'

def load_watchlist_from_gist():
    try:
        r = requests.get(
            f'https://api.github.com/gists/{GITHUB_GIST_ID}',
            headers={'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github+json'},
            timeout=10
        )
        r.raise_for_status()
        content = r.json()['files'][GIST_FILENAME]['content']
        tickers = json.loads(content)
        return tickers if isinstance(tickers, list) else None
    except Exception as e:
        print('Gist load failed:', e)
        return None

def save_watchlist_to_gist(tickers):
    try:
        r = requests.patch(
            f'https://api.github.com/gists/{GITHUB_GIST_ID}',
            headers={'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github+json'},
            json={'files': {GIST_FILENAME: {'content': json.dumps(tickers)}}},
            timeout=10
        )
        r.raise_for_status()
        return True
    except Exception as e:
        print('Gist save failed:', e)
        return False

def load_anchors_from_gist():
    """Per-ticker Stage 1 anchor state: {sym: {'anchor': low_price, 'date': 'YYYY-MM-DD'}}"""
    try:
        r = requests.get(
            f'https://api.github.com/gists/{GITHUB_GIST_ID}',
            headers={'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github+json'},
            timeout=10
        )
        r.raise_for_status()
        files = r.json().get('files', {})
        if ANCHOR_FILENAME not in files:
            return {}
        content = files[ANCHOR_FILENAME]['content']
        anchors = json.loads(content)
        return anchors if isinstance(anchors, dict) else {}
    except Exception as e:
        print('Anchor load failed:', e)
        return {}

def save_anchors_to_gist(anchors):
    try:
        r = requests.patch(
            f'https://api.github.com/gists/{GITHUB_GIST_ID}',
            headers={'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github+json'},
            json={'files': {ANCHOR_FILENAME: {'content': json.dumps(anchors)}}},
            timeout=10
        )
        r.raise_for_status()
        return True
    except Exception as e:
        print('Anchor save failed:', e)
        return False

POSITIONS_FILENAME = 'golden_scanner_positions.json'

def load_positions_from_gist():
    """{'active': {sym: {shares, avg_price, realized_pl, buys:[...], sells:[...]}}, 'closed': [...]}"""
    try:
        r = requests.get(
            f'https://api.github.com/gists/{GITHUB_GIST_ID}',
            headers={'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github+json'},
            timeout=10
        )
        r.raise_for_status()
        files = r.json().get('files', {})
        if POSITIONS_FILENAME not in files:
            return {'active': {}, 'closed': []}
        content = files[POSITIONS_FILENAME]['content']
        data = json.loads(content)
        if not isinstance(data, dict): return {'active': {}, 'closed': []}
        data.setdefault('active', {}); data.setdefault('closed', [])
        return data
    except Exception as e:
        print('Positions load failed:', e)
        return {'active': {}, 'closed': []}

def save_positions_to_gist(data):
    try:
        r = requests.patch(
            f'https://api.github.com/gists/{GITHUB_GIST_ID}',
            headers={'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github+json'},
            json={'files': {POSITIONS_FILENAME: {'content': json.dumps(data)}}},
            timeout=10
        )
        r.raise_for_status()
        return True
    except Exception as e:
        print('Positions save failed:', e)
        return False

# ── DEFAULT WATCHLIST ────────────────────────────────────────────────────────
DEFAULT_WATCHLIST = [
    "AAL","AAOI","AAPL","ADBE","AEHR","AGQ","AI","AIR","ALLY","ALT",
    "AMD","AMGN","AMKR","AMZN","ARM","ASML","ASTS","AVAV","AVGO",
    "B","BA","BABA","BABX","BAC","BAB","BETZ","BITO","BLK","BLUE",
    "BMNR","BOIL","BORR","BOTZ","BX","BXMT",
    "C","CART","CAT","CCL","CELH","CIFR","CL","CLF","CLOU","CLSK",
    "COHR","COIN","CONL","COP","COPX","COST","CRCL","CRDO","CRM",
    "CRWD","CRWV","CSCO","CURE","CVS","CVX",
    "D","DAL","DDOG","DE","DELL","DFEN","DG","DIS","DJT","DKNG",
    "DNUT","DOW","DPST","DUK","DVN",
    "ELF","ENVX","EOSE","ERX","ESLT","ET","ETHU",
    "F","FCX","FDX","FE","FLR","FUBO",
    "GENI","GEV","GILD","GLD","GLW","GOOGL","GS","GUSH",
    "HAL","HAS","HE","HERO","HIMS","HOOD","HRZN","HSY","HUM",
    "ICLN","IETC","ILMN","INTC","IONQ","IREN","ISRG",
    "JACK","JNJ","JOBY","JPM",
    "K","KHC","KLAR","KLG","KO","KTOS","KVUE",
    "LDOS","LEU","LITE","LMND","LLY","LMT","LOW","LULU","LYFT",
    "MARA","MCD","MDB","MDLN","MDLZ","META","METV","MLM","MMM",
    "MO","MP","MPT","MRK","MRNA","MRVL","MSFT","MSTR","MSTY","MU",
    "NAIL","NBIS","NET","NFLX","NEM","NIO","NKE","NOW","NU","NUE",
    "NUGT","NVDA","NVO",
    "O","OARK","OHI","OILU","OXY",
    "PANW","PCG","PEP","PFE","PG","PILL","PINS","PLTR","PLUG",
    "PNNT","PSX","PYPL","PZZA",
    "QBTS","QQQ","QQQM",
    "RBLX","RDDT","RDW","RIOT","RIVN","ROBO","ROKU","RTX","RUN","RZLT",
    "SBUX","SHOP","SLV","SMH","SNAP","SNOW","SNY","SO","SOFX","SOUN",
    "SOXL","SPCX","SPXL","SPXS","SPXU","SPYM","SQQQ","SRAD","STM","STZ","SYS",
    "T","TDV","TEAM","TGT","TLRY","TMUS","TREE","TSLA","TSLL","TSM","TTD",
    "UA","UBER","UBRL","ULTA","UMAC","UNH","UNHG","UP","URA",
    "URAA","URAX","URI","USO","UTSL","UX",
    "VALE","VFC","VIAV","VMC","VST","VZ",
    "WBA","WBD","WDAY","WELL","WEN","WFC","WM","WMT",
    "X","XOM","XOVR","XSD","XT","XYZ",
    "ZETA","ZIM","ZS",
    "EQT","KMI","WMB","BE","CMI","ETN","PWR","VRT","HUBB","NVT","EMR",
    "OKLO","SMR","NNE","CEG","BWXT","UUUU",
    "WULF","APLD",
    "RKLB","LUNR","GSAT","BKSY","SPIR","PL",
    "AA","ATI","TECK","HXL","CRS","PKE","MTRN","QRVO","ADI","FLY",
    "AMAT","KLAC","LRCX","ALAB","ANET","AXTI","RBRK","TEM",
    "NOC","RCAT","DPRO","AVEX","ONDS",
    "LAC","TMQ",
]

# ── STAGE SCANNER ─────────────────────────────────────────────────────────────
def ema(series, period):
    return series.ewm(span=period, adjust=False).mean()

def calc_tsi(closes, long=25, short=13, signal=8):
    pc=closes.diff(); apc=pc.abs()
    e1=ema(pc,long); e2=ema(e1,short)
    ae1=ema(apc,long); ae2=ema(ae1,short)
    tsi=(e2/ae2.replace(0,float('nan')))*100
    return tsi, ema(tsi,signal)

def fetch_ticker(sym):
    try:
        tk=yf.Ticker(sym)
        daily=tk.history(period="2y",interval="1d")
        if daily.empty or len(daily)<60: return None
        weekly=tk.history(period="5y",interval="1wk")
        if weekly.empty or len(weekly)<60: weekly=None
        return daily,weekly
    except: return None

def classify_ticker(sym,daily,weekly,anchors):
    try:
        import math
        closes_d=daily['Close'].dropna()
        if len(closes_d)<60: return None
        lows_d=daily['Low'].reindex(closes_d.index)
        price=float(closes_d.iloc[-1]); prev=float(closes_d.iloc[-2])
        if math.isnan(price) or math.isnan(prev): return None
        tsi_d,_=calc_tsi(closes_d)
        ema7_d=ema(closes_d,7); ema21_d=ema(closes_d,21)
        sma200_d=closes_d.rolling(200).mean()
        tsi_val=float(tsi_d.iloc[-1]); tsi_prev=float(tsi_d.iloc[-2])
        e7=float(ema7_d.iloc[-1]); e21=float(ema21_d.iloc[-1])
        s200=float(sma200_d.iloc[-1]) if not pd.isna(sma200_d.iloc[-1]) else None
        prev_e21=float(ema21_d.iloc[-2])
        ab7=price>e7; ab21=price>e21
        ab200=(price>s200) if s200 else None
        prev_ab21=prev>prev_e21
        tsi_dir="↑" if tsi_val>tsi_prev else "↓"
        wtsi_val=None
        if weekly is not None and not weekly.empty:
            cw=weekly['Close'].dropna()
            if len(cw)>=60:
                tw,_=calc_tsi(cw); wtsi_val=float(tw.iloc[-1])

        # ── ANCHOR STATE ── (persists across scans via Gist)
        entry=anchors.get(sym)
        anchor_price=entry['anchor'] if entry else None
        anchor_holding=(anchor_price is not None and price>=anchor_price)
        is_trend_anchor=bool(entry and entry.get('trend'))
        s1_label='🟢 STAGE 1 (TREND)' if is_trend_anchor else '🟢 STAGE 1'
        cont_label='🔵 STAGE 1 CONT. (TREND)' if is_trend_anchor else '🔵 STAGE 1 CONT.'

        if not ab21 and not prev_ab21 and price<prev and tsi_val>-20:
            if anchor_holding:
                sk='stage-1-green'; label=s1_label; buy=True
                signal=f'TSI {tsi_val:.1f}. Pulled back below 21 EMA but anchor ${anchor_price:.2f} still holding. Still accumulating.'
            else:
                sk='stage-4'; label='🔴 STAGE 4'; buy=False
                signal='Two candles below 21 EMA. Price declining. EXIT position.'
        elif not ab7 and ab21:
            if anchor_holding:
                sk='stage-1-blue'; label=cont_label; buy=True
                signal=f'TSI {tsi_val:.1f}. Pulled back below 7 EMA but above anchor ${anchor_price:.2f}. Still accumulating.'
            else:
                sk='stage-3'; label='🟠 STAGE 3'; buy=False
                if tsi_val>40: signal=f'TSI {tsi_val:.1f} — Take HEAVY profits now.'
                elif tsi_val>20: signal=f'TSI {tsi_val:.1f} — Start taking small profits.'
                else: signal=f'TSI {tsi_val:.1f}. Distribution zone. Watch closely.'
        elif ab21 and not prev_ab21:
            sk='stage-2b'; label='🟣 S2 BREAKOUT'; buy=True
            signal='Candle broke above 21 EMA. Last buy before the run.'
        elif not ab21:
            if tsi_val<-20 and ab7:
                if not entry:
                    # Anchor = lowest wick across the WHOLE basing period where TSI
                    # has stayed continuously below -20, up through today's
                    # confirmation candle — not just today's low.
                    low_min=float(lows_d.iloc[-1])
                    j=len(tsi_d)-2
                    while j>=0 and float(tsi_d.iloc[j])<-20:
                        lj=lows_d.iloc[j]
                        if not pd.isna(lj): low_min=min(low_min,float(lj))
                        j-=1
                    anchors[sym]={'anchor':low_min,'date':date.today().strftime("%Y-%m-%d"),'trend':False}
                if wtsi_val and wtsi_val<-20:
                    sk='gold'; label='🥇 GOLD SETUP'; buy=True
                    signal=f'{"🔥 TSI "+str(round(tsi_val,1))+" EXTREME. " if tsi_val<-40 else "TSI "+str(round(tsi_val,1))+". "}Weekly {wtsi_val:.1f}. HEAVY BUY.'
                else:
                    sk='stage-1-green'; label='🟢 STAGE 1'; buy=True
                    signal=f'{"🔥 TSI "+str(round(tsi_val,1))+" EXTREMELY HOT. HEAVY BUY." if tsi_val<-40 else "TSI "+str(round(tsi_val,1))+". Candle above 7 EMA. Accumulate."}'
            elif ab200 and tsi_val<5 and ab7 and not entry:
                # Trend-based Stage 1: strong long-term uptrend (above 200 SMA),
                # TSI dipped under 5 without ever reaching -20, and price already
                # confirmed by closing back above the 7 EMA. Fires on confirmation
                # only — no separate "Watch" state for this narrower threshold.
                low_min=float(lows_d.iloc[-1])
                j=len(tsi_d)-2
                while j>=0 and float(tsi_d.iloc[j])<5:
                    lj=lows_d.iloc[j]
                    if not pd.isna(lj): low_min=min(low_min,float(lj))
                    j-=1
                anchors[sym]={'anchor':low_min,'date':date.today().strftime("%Y-%m-%d"),'trend':True}
                sk='stage-1-green'; label='🟢 STAGE 1 (TREND)'; buy=True
                signal=f'TSI {tsi_val:.1f} (200 SMA uptrend, never hit -20). Candle above 7 EMA. Accumulate.'
            elif tsi_val<-20 and not ab7:
                if entry and not anchor_holding:
                    sk='stage-1-yellow'; label='🟡 WATCH'; buy=False
                    signal=f'⚠ Anchor ${anchor_price:.2f} broken. TSI {tsi_val:.1f}. Waiting for a new Stage 1 confirmation.'
                    anchors.pop(sym,None)
                else:
                    sk='stage-1-yellow'; label='🟡 WATCH'; buy=False
                    signal=f'{"🔥 " if tsi_val<-40 else ""}TSI {tsi_val:.1f}. Wait for candle close above 7 EMA.'
            elif -20<=tsi_val<=0 and ab7:
                sk='stage-1-blue'; label=cont_label; buy=True
                signal=f'TSI {tsi_val:.1f} (below 0). Candle above 7 EMA. Accumulate lightly.'
            else:
                sk='stage-2'; label='⬜ STAGE 2'; buy=False; signal='Running. Hold.'
        else:
            sk='stage-2'; label='⬜ STAGE 2'; buy=False; signal='Running above 7 & 21 EMA. Hold.'

        # Cycle retires once TSI reaches the sell zone — anchor no longer relevant
        if entry and tsi_val>=20:
            anchors.pop(sym,None)

        return {
            'sym':sym,'price':round(price,2),'tsi':round(tsi_val,2),
            'tsi_dir':tsi_dir,'wtsi':round(wtsi_val,2) if wtsi_val else None,
            'ab7':ab7,'ab21':ab21,'ab200':ab200,
            'sk':sk,'label':label,'signal':signal,'buy':buy,'is_gold':sk=='gold',
        }
    except: return None

def tsi_zone(v):
    if v is None: return '—'
    if v<-40: return '🔥🔥 EXTREME HOT'
    if v<-20: return '🔥 BUY ZONE'
    if v>40: return '⚠ HEAVY PROFITS'
    if v>20: return '⚠ TAKE PROFITS'
    if v<0: return 'BELOW ZERO'
    return 'NEUTRAL'

# ── SCAN STATE ────────────────────────────────────────────────────────────────
scan_state = {'running':False,'progress':0,'total':0,'current':'',
              'results':[],'errors':[],'done':False,'date':''}

def run_scan_thread(watchlist):
    global scan_state
    scan_state.update({'running':True,'done':False,'results':[],'errors':[],
                       'total':len(watchlist),'progress':0,
                       'date':date.today().strftime("%B %d, %Y")})
    anchors=load_anchors_from_gist()
    anchors_before=json.dumps(anchors,sort_keys=True)
    for i,sym in enumerate(watchlist):
        scan_state['current']=sym; scan_state['progress']=i+1
        data=fetch_ticker(sym)
        if data is None: scan_state['errors'].append(sym); continue
        result=classify_ticker(sym,data[0],data[1],anchors)
        if result: scan_state['results'].append(result)
        else: scan_state['errors'].append(sym)
    if json.dumps(anchors,sort_keys=True)!=anchors_before:
        save_anchors_to_gist(anchors)
    scan_state['running']=False; scan_state['done']=True

# ── OPTION CHAIN STATE ────────────────────────────────────────────────────────
option_jobs = {}  # sym -> {state, result, progress}

def run_option_thread(sym):
    from option_engine import full_analysis
    option_jobs[sym] = {'state':'running','result':None,'progress':0,'total':0,'exp':''}
    def cb(i, total, exp):
        option_jobs[sym]['progress'] = i
        option_jobs[sym]['total']    = total
        option_jobs[sym]['exp']      = exp
    try:
        result = full_analysis(sym, progress_cb=cb)
        option_jobs[sym]['result'] = result
        option_jobs[sym]['state']  = 'done'
    except Exception as e:
        option_jobs[sym]['state'] = 'error'
        option_jobs[sym]['error'] = str(e)

# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────
def build_excel(results, scan_date):
    def fill(c): return PatternFill("solid",start_color=c)
    def font(c,bold=False,sz=10): return Font(name="Calibri",color=c,bold=bold,size=sz)
    RF={'gold':fill("1C1400"),'stage-1-green':fill("0A1F0A"),'stage-1-yellow':fill("1F1A00"),
        'stage-1-blue':fill("051520"),'stage-2b':fill("120A20"),'stage-3':fill("1F0E00"),
        'stage-4':fill("1F0505"),'stage-2':fill("0D1117")}
    RFA={'gold':fill("221900"),'stage-1-green':fill("0D2410"),'stage-1-yellow':fill("231E00"),
         'stage-1-blue':fill("071A28"),'stage-2b':fill("160D25"),'stage-3':fill("251200"),
         'stage-4':fill("250707"),'stage-2':fill("161B22")}
    LC={'gold':"F0A500",'stage-1-green':"00CC66",'stage-1-yellow':"FFD700",
        'stage-1-blue':"66CCFF",'stage-2b':"CC88FF",'stage-3':"FF8C00",
        'stage-4':"FF4444",'stage-2':"8B949E"}
    SN={'gold':"🥇 GOLD SETUPS","stage-1-green":"🟢 STAGE 1 CONFIRMED",
        'stage-1-yellow':"🟡 WATCHING",'stage-1-blue':"🔵 STAGE 1 CONTINUATION",
        'stage-2b':"🟣 S2 BREAKOUT",'stage-3':"🟠 STAGE 3",'stage-4':"🔴 STAGE 4 — EXIT"}
    thin=Side(style="thin",color="21262D")
    BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
    C=Alignment(horizontal="center",vertical="center")
    L=Alignment(horizontal="left",vertical="center",wrap_text=True)
    MUT=font("8B949E"); WH=font("E6EDF3"); UP=font("00CC66"); DN=font("FF4444")
    HDR=Font(name="Calibri",color="6E7681",bold=True,size=9)
    def tf(v):
        if v is None: return MUT
        if v<-40: return font("FF6B6B",True)
        if v<-20: return font("00CC66",True)
        if v<0: return font("66CCFF")
        if v>40: return font("FF4444",True)
        if v>20: return font("FF8C00",True)
        return font("8B949E")
    HEADERS=["TICKER","PRICE","STAGE","DAILY TSI","DIR","TSI ZONE","WEEKLY TSI","7 EMA","21 EMA","200 SMA","BUY?","ACTION / SIGNAL"]
    WIDTHS=[11,9,20,11,7,18,12,8,8,10,7,58]
    ORDER={'gold':0,'stage-1-green':1,'stage-1-yellow':2,'stage-1-blue':3,'stage-2b':4,'stage-3':5,'stage-4':6,'stage-2':7}
    actionable=[r for r in results if r['sk']!='stage-2']
    actionable.sort(key=lambda r:(ORDER.get(r['sk'],9),r['sym']))
    def build_sheet(wb,title,tab_color,data,is_first=False):
        ws=wb.active if is_first else wb.create_sheet(title)
        if is_first: ws.title=title
        ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=tab_color
        for row in ws.iter_rows(min_row=1,max_row=max(len(data)+20,30),min_col=1,max_col=12):
            for cell in row: cell.fill=fill("06080B")
        ws.merge_cells("A1:L1"); ws["A1"]="✦ GOLDEN SCANNER — Gerald Peters Stage System"
        ws["A1"].font=Font(name="Calibri",bold=True,color="F0A500",size=14)
        ws["A1"].fill=fill("0A0D12"); ws["A1"].alignment=L; ws.row_dimensions[1].height=28
        ws.merge_cells("A2:L2")
        ws["A2"]=f"TSI 25/13/8 EMA · Daily + Weekly · Yahoo Finance · {scan_date} · {len(results)} tickers"
        ws["A2"].font=Font(name="Calibri",color="8B949E",size=9)
        ws["A2"].fill=fill("0A0D12"); ws["A2"].alignment=L; ws.row_dimensions[2].height=14
        ws.merge_cells("A3:L3")
        ws["A3"]="🥇 GOLD  🟢 Stage 1  🟡 Watch  🔵 Cont.  🟣 S2 Break  🟠 Stage 3  🔴 Stage 4  · ↑ curling up  ↓ falling"
        ws["A3"].font=Font(name="Calibri",color="6E7681",size=8,italic=True)
        ws["A3"].fill=fill("0A0D12"); ws["A3"].alignment=L; ws.row_dimensions[3].height=13
        ws.row_dimensions[4].height=5
        for ci,(h,w) in enumerate(zip(HEADERS,WIDTHS),1):
            cell=ws.cell(row=5,column=ci,value=h)
            cell.font=HDR; cell.fill=fill("0D1117"); cell.alignment=C; cell.border=BORDER
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.row_dimensions[5].height=20
        if not data:
            ws.merge_cells("A6:L6"); ws["A6"]="No stocks in this category right now."
            ws["A6"].font=font("8B949E",False,10); ws["A6"].fill=fill("06080B")
            ws["A6"].alignment=C; ws.row_dimensions[6].height=30
            ws.freeze_panes="A6"; return ws
        rn=6; lsk="___"; rc=0
        for r in data:
            sk=r['sk']; lc=LC.get(sk,"8B949E")
            if sk!=lsk:
                ws.merge_cells(f"A{rn}:L{rn}"); ws[f"A{rn}"]=SN.get(sk,'')
                ws[f"A{rn}"].font=Font(name="Calibri",bold=True,color=lc,size=9,italic=True)
                ws[f"A{rn}"].fill=fill("0D1117"); ws[f"A{rn}"].alignment=L
                ws.row_dimensions[rn].height=16; rn+=1; lsk=sk; rc=0
            rf=RF.get(sk,fill("0D1117")) if rc%2==0 else RFA.get(sk,fill("161B22")); rc+=1
            p=r['price']; tsi=r['tsi']; wtsi=r['wtsi']; td=r.get('tsi_dir','—'); ab200=r['ab200']
            vals=[r['sym'],f"${p:.2f}" if p else '—',r['label'],
                  f"{tsi:.2f}" if tsi is not None else '—',td,tsi_zone(tsi),
                  f"{wtsi:.2f}" if wtsi is not None else '—',
                  '✓' if r['ab7'] else '✗','✓' if r['ab21'] else '✗',
                  ('✓' if ab200 else '✗') if ab200 is not None else '—',
                  '✓ BUY' if r['buy'] else '—',r['signal']]
            rfs=[Font(name="Calibri",color=lc,bold=True,size=11),WH,
                 Font(name="Calibri",color=lc,bold=True,size=10),
                 tf(tsi),font("00CC66",True) if td=="↑" else font("FF4444",True),tf(tsi),tf(wtsi),
                 UP if r['ab7'] else DN,UP if r['ab21'] else DN,
                 (UP if ab200 else DN) if ab200 is not None else MUT,
                 Font(name="Calibri",color="00CC66",bold=True) if r['buy'] else MUT,WH]
            als=[C]*11+[L]
            for ci,(val,fnt,aln) in enumerate(zip(vals,rfs,als),1):
                cell=ws.cell(row=rn,column=ci,value=val)
                cell.font=fnt; cell.fill=rf; cell.alignment=aln; cell.border=BORDER
            ws.row_dimensions[rn].height=18; rn+=1
        ws.freeze_panes="A6"
    wb=Workbook()
    build_sheet(wb,"Golden Scanner","F0A500",actionable,is_first=True)
    build_sheet(wb,"⚡ ACTION REQUIRED","00CC66",[r for r in actionable if r['buy']])
    build_sheet(wb,"🟢 BUY SIGNALS","00CC66",[r for r in actionable if r['sk'] in ('gold','stage-1-green','stage-1-blue','stage-2b')])
    build_sheet(wb,"🟡 WATCHING","FFD700",[r for r in actionable if r['sk']=='stage-1-yellow'])
    build_sheet(wb,"🟠 TAKE PROFITS","FF8C00",[r for r in actionable if r['sk']=='stage-3'])
    build_sheet(wb,"🔴 EXIT NOW","FF4444",[r for r in actionable if r['sk']=='stage-4'])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/scan', methods=['POST'])
def start_scan():
    global scan_state
    if scan_state['running']: return jsonify({'error':'Scan already running'}),400
    data=request.json
    watchlist=data.get('watchlist',DEFAULT_WATCHLIST)
    watchlist=[t.strip().upper() for t in watchlist if t.strip()]
    t=threading.Thread(target=run_scan_thread,args=(watchlist,))
    t.daemon=True; t.start()
    return jsonify({'status':'started','total':len(watchlist)})

@app.route('/api/progress')
def get_progress():
    s=scan_state
    counts={'gold':0,'stage-1-green':0,'stage-1-yellow':0,'stage-1-blue':0,
            'stage-2b':0,'stage-3':0,'stage-4':0,'stage-2':0}
    for r in s['results']: counts[r['sk']]=counts.get(r['sk'],0)+1
    return jsonify({'running':s['running'],'done':s['done'],'progress':s['progress'],
                    'total':s['total'],'current':s['current'],'counts':counts,
                    'scanned':len(s['results']),'errors':len(s['errors']),'date':s['date']})

@app.route('/api/results')
def get_results():
    ORDER={'gold':0,'stage-1-green':1,'stage-1-yellow':2,'stage-1-blue':3,
           'stage-2b':4,'stage-3':5,'stage-4':6,'stage-2':7}
    results=sorted(scan_state['results'],key=lambda r:(ORDER.get(r['sk'],9),r['sym']))
    return jsonify({'results':results,'errors':scan_state['errors'],'date':scan_state['date']})

@app.route('/api/download')
def download_excel():
    if not scan_state['results']: return jsonify({'error':'No results yet'}),400
    buf=build_excel(scan_state['results'],scan_state['date'])
    filename=f"Golden_Scanner_{scan_state['date'].replace(' ','_').replace(',','')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/watchlist', methods=['GET','POST'])
def watchlist_api():
    wl_file='watchlist.json'
    if request.method=='POST':
        data=request.json
        tickers=[t.strip().upper() for t in data.get('tickers',[]) if t.strip()]
        synced=save_watchlist_to_gist(tickers)
        try:
            with open(wl_file,'w') as f: json.dump(tickers,f)
        except Exception:
            pass
        return jsonify({'saved':True,'count':len(tickers),'synced':synced})
    # GET — Gist is the source of truth so every device sees the same list
    tickers=load_watchlist_from_gist()
    if tickers is None:
        if os.path.exists(wl_file):
            with open(wl_file) as f: tickers=json.load(f)
        else:
            tickers=DEFAULT_WATCHLIST
    return jsonify({'tickers':tickers,'count':len(tickers)})

@app.route('/api/positions', methods=['GET'])
def positions_api():
    return jsonify(load_positions_from_gist())

@app.route('/api/positions/buy', methods=['POST'])
def positions_buy():
    data=request.json
    sym=data.get('sym','').strip().upper()
    price=float(data.get('price')); shares=float(data.get('shares'))
    if not sym or price<=0 or shares<=0:
        return jsonify({'error':'Invalid input'}),400
    pos=load_positions_from_gist()
    today=date.today().strftime("%Y-%m-%d")
    entry=pos['active'].get(sym)
    buy_record={'date':today,'price':round(price,4),'shares':round(shares,6),'amount':round(price*shares,2)}
    if entry:
        old_shares=entry['shares']; old_avg=entry['avg_price']
        new_shares=old_shares+shares
        new_avg=((old_shares*old_avg)+(shares*price))/new_shares
        entry['shares']=round(new_shares,6); entry['avg_price']=round(new_avg,4)
        entry['buys'].append(buy_record)
    else:
        entry={'shares':round(shares,6),'avg_price':round(price,4),'realized_pl':0.0,
               'buys':[buy_record],'sells':[]}
        pos['active'][sym]=entry
    synced=save_positions_to_gist(pos)
    return jsonify({'saved':True,'synced':synced,'position':entry})

@app.route('/api/positions/delete_sell', methods=['POST'])
def positions_delete_sell():
    data=request.json
    sym=data.get('sym','').strip().upper()
    sell_index=data.get('sell_index')
    closed_index=data.get('closed_index')  # present only when deleting from a closed position
    pos=load_positions_from_gist()

    if closed_index is None:
        entry=pos['active'].get(sym)
        if not entry or sell_index is None or sell_index>=len(entry['sells']):
            return jsonify({'error':'Sell not found'}),400
        removed=entry['sells'].pop(sell_index)
        entry['shares']=round(entry['shares']+removed['shares'],6)
        entry['realized_pl']=round(entry.get('realized_pl',0.0)-removed['realized_pl'],2)
        synced=save_positions_to_gist(pos)
        return jsonify({'saved':True,'synced':synced,'reopened':False})

    closed_index=int(closed_index)
    if closed_index<0 or closed_index>=len(pos['closed']):
        return jsonify({'error':'Closed position not found'}),400
    c=pos['closed'][closed_index]
    if sell_index is None or sell_index>=len(c['sells']):
        return jsonify({'error':'Sell not found'}),400
    removed=c['sells'].pop(sell_index)
    remaining_sold=sum(s['shares'] for s in c['sells'])
    remaining_shares=round(c['total_shares']-remaining_sold,6)

    if remaining_shares<=0.0001:
        # Still fully closed even without that sell — recompute totals
        c['total_proceeds']=round(sum(s['amount'] for s in c['sells']),2)
        c['realized_pl']=round(c['total_proceeds']-c['total_invested'],2)
        synced=save_positions_to_gist(pos)
        return jsonify({'saved':True,'synced':synced,'reopened':False})

    # Deleting that sell leaves shares open again — reopen as an active position
    pos['closed'].pop(closed_index)
    realized_remaining=round(sum(s['realized_pl'] for s in c['sells']),2)
    existing=pos['active'].get(sym)
    if existing:
        old_shares=existing['shares']; old_avg=existing['avg_price']
        new_shares=round(old_shares+remaining_shares,6)
        new_avg=((old_shares*old_avg)+(remaining_shares*c['avg_buy_price']))/new_shares if new_shares>0 else c['avg_buy_price']
        existing['shares']=new_shares; existing['avg_price']=round(new_avg,4)
        existing['buys']=existing['buys']+c['buys']
        existing['sells']=existing['sells']+c['sells']
        existing['realized_pl']=round(existing.get('realized_pl',0.0)+realized_remaining,2)
    else:
        pos['active'][sym]={'shares':remaining_shares,'avg_price':c['avg_buy_price'],
                             'realized_pl':realized_remaining,'buys':c['buys'],'sells':c['sells']}
    synced=save_positions_to_gist(pos)
    return jsonify({'saved':True,'synced':synced,'reopened':True})

@app.route('/api/positions/sell', methods=['POST'])
def positions_sell():
    data=request.json
    sym=data.get('sym','').strip().upper()
    price=float(data.get('price')); shares=float(data.get('shares'))
    if not sym or price<=0 or shares<=0:
        return jsonify({'error':'Invalid input'}),400
    pos=load_positions_from_gist()
    entry=pos['active'].get(sym)
    if not entry:
        return jsonify({'error':f'No open position in {sym}'}),400
    shares=min(shares,entry['shares'])  # can't sell more than owned
    today=date.today().strftime("%Y-%m-%d")
    realized=(price-entry['avg_price'])*shares
    sell_record={'date':today,'price':round(price,4),'shares':round(shares,6),
                 'amount':round(price*shares,2),'realized_pl':round(realized,2)}
    entry['sells'].append(sell_record)
    entry['realized_pl']=round(entry.get('realized_pl',0.0)+realized,2)
    entry['shares']=round(entry['shares']-shares,6)
    closed_entry=None
    if entry['shares']<=0.0001:
        total_bought=sum(b['shares'] for b in entry['buys'])
        total_invested=round(total_bought*entry['avg_price'],2)
        total_proceeds=round(sum(s['amount'] for s in entry['sells']),2)
        closed_entry={'ticker':sym,'closed_date':today,'total_shares':round(total_bought,6),
                      'avg_buy_price':entry['avg_price'],'total_invested':total_invested,
                      'total_proceeds':total_proceeds,
                      'realized_pl':round(total_proceeds-total_invested,2),
                      'buys':entry['buys'],'sells':entry['sells']}
        pos['closed'].append(closed_entry)
        del pos['active'][sym]
    synced=save_positions_to_gist(pos)
    return jsonify({'saved':True,'synced':synced,'closed':closed_entry is not None,
                    'position':closed_entry if closed_entry else entry})

# ── OPTION CHAIN ROUTES ───────────────────────────────────────────────────────
@app.route('/api/options/start/<sym>', methods=['POST'])
def start_option_analysis(sym):
    sym=sym.upper()
    if sym in option_jobs and option_jobs[sym].get('state')=='running':
        return jsonify({'status':'already_running'})
    t=threading.Thread(target=run_option_thread,args=(sym,))
    t.daemon=True; t.start()
    return jsonify({'status':'started','symbol':sym})

@app.route('/api/options/progress/<sym>')
def option_progress(sym):
    sym=sym.upper()
    job=option_jobs.get(sym,{})
    return jsonify({
        'state':    job.get('state','not_started'),
        'progress': job.get('progress',0),
        'total':    job.get('total',0),
        'exp':      job.get('exp',''),
        'error':    job.get('error',''),
    })

@app.route('/api/options/result/<sym>')
def option_result(sym):
    sym=sym.upper()
    job=option_jobs.get(sym,{})
    if job.get('state')!='done':
        return jsonify({'state':job.get('state','not_started')}),202
    result=job['result']
    if not result: return jsonify({'error':'No data'}),404
    # Don't send full chart_data here — send via separate endpoint
    r={k:v for k,v in result.items() if k!='chart_data'}
    r['state']='done'
    return jsonify(r)

@app.route('/api/options/chart/<sym>')
def option_chart(sym):
    sym=sym.upper()
    job=option_jobs.get(sym,{})
    if job.get('state')!='done': return jsonify({'error':'Not ready'}),202
    result=job['result']
    if not result: return jsonify({'error':'No data'}),404
    return jsonify({'chart_data':result['chart_data'],'spot':result['spot']})

@app.route('/api/options/pine/<sym>')
def option_pine_single(sym):
    sym=sym.upper()
    job=option_jobs.get(sym,{})
    if job.get('state')!='done': return jsonify({'error':'Not ready'}),202
    result=job['result']
    if not result: return jsonify({'error':'No data'}),404
    from option_engine import generate_pine_single
    pine=generate_pine_single(result)
    return jsonify({'pine':pine,'symbol':sym})

@app.route('/api/options/pine/universal', methods=['POST'])
def option_pine_universal():
    from option_engine import generate_pine_universal
    syms=request.json.get('symbols',[])
    analyses=[]
    for sym in syms:
        sym=sym.upper()
        job=option_jobs.get(sym,{})
        if job.get('state')=='done' and job.get('result'):
            analyses.append(job['result'])
    if not analyses: return jsonify({'error':'No analyzed symbols ready'}),400
    pine=generate_pine_universal(analyses)
    return jsonify({'pine':pine,'count':len(analyses)})

@app.route('/api/options/export/<sym>')
def export_chain(sym):
    """Export full option chain — ALL expirations, ALL strikes, ALL greeks to Excel"""
    import requests as req, time as t
    from datetime import date, datetime
    sym = sym.upper()

    TRADIER_TOKEN = "Vzj4eByTvakHHT4PLg4jl6AjtfZs"
    BASE  = "https://api.tradier.com/v1"
    HDR   = {"Authorization": f"Bearer {TRADIER_TOKEN}", "Accept": "application/json"}

    def fill(c): return PatternFill("solid", start_color=c)
    def font(c, bold=False, sz=9): return Font(name="Calibri", color=c, bold=bold, size=sz)

    try:
        # Spot price
        r = req.get(f"{BASE}/markets/quotes", params={"symbols": sym}, headers=HDR, timeout=10)
        q = r.json()["quotes"]["quote"]
        spot = float(q.get("last") or q.get("close") or 0)

        # All expirations
        r = req.get(f"{BASE}/markets/options/expirations",
                    params={"symbol": sym, "includeAllRoots": "true"},
                    headers=HDR, timeout=10)
        exps = r.json()["expirations"]["date"]
        if isinstance(exps, str): exps = [exps]

        wb = Workbook()
        wb.remove(wb.active)

        # ── SUMMARY SHEET ──────────────────────────────────────────────────
        ws_sum = wb.create_sheet("SUMMARY", 0)
        ws_sum.sheet_properties.tabColor = "F0A500"
        ws_sum.sheet_view.showGridLines = False

        for row in ws_sum.iter_rows(min_row=1, max_row=len(exps)+10, min_col=1, max_col=10):
            for cell in row: cell.fill = fill("06080B")

        ws_sum.merge_cells("A1:J1")
        ws_sum.cell(1,1, f"✦ {sym} — Full Option Chain Export | Spot: ${spot:.2f} | {datetime.now().strftime('%Y-%m-%d %H:%M')} | All Expirations")
        ws_sum.cell(1,1).font = Font(name="Calibri", bold=True, color="F0A500", size=13)
        ws_sum.cell(1,1).fill = fill("0A0D12")
        ws_sum.row_dimensions[1].height = 24

        sum_hdrs = ["EXPIRATION","DAYS OUT","TOTAL STRIKES","CALLS","PUTS","TOTAL OI","CALL OI","PUT OI","AVG IV","NET DELTA"]
        sum_widths = [14,10,14,10,10,12,10,10,10,12]
        for ci,(h,w) in enumerate(zip(sum_hdrs,sum_widths),1):
            cell = ws_sum.cell(3, ci, h)
            cell.font = Font(name="Calibri", bold=True, color="6E7681", size=8)
            cell.fill = fill("0D1117")
            from openpyxl.utils import get_column_letter
            ws_sum.column_dimensions[get_column_letter(ci)].width = w
        ws_sum.row_dimensions[3].height = 18

        sum_row = 4

        # ── PER EXPIRATION ──────────────────────────────────────────────────
        for exp in exps:
            try:
                r2 = req.get(f"{BASE}/markets/options/chains",
                             params={"symbol": sym, "expiration": exp, "greeks": "true"},
                             headers=HDR, timeout=15)
                data = r2.json().get("options", {})
                t.sleep(0.1)

                if not data or not data.get("option"):
                    continue

                opts = data["option"]
                if isinstance(opts, dict): opts = [opts]

                # Sort by strike then type
                opts_sorted = sorted(opts, key=lambda o: (
                    float(o.get("strike", 0)),
                    0 if o.get("option_type","").lower()=="call" else 1
                ))

                calls = [o for o in opts_sorted if o.get("option_type","").lower()=="call"]
                puts  = [o for o in opts_sorted if o.get("option_type","").lower()=="put"]
                strikes = sorted(set(float(o.get("strike",0)) for o in opts_sorted))

                total_oi   = sum(int(o.get("open_interest") or 0) for o in opts_sorted)
                call_oi    = sum(int(o.get("open_interest") or 0) for o in calls)
                put_oi     = sum(int(o.get("open_interest") or 0) for o in puts)
                ivs        = [float(o.get("greeks",{}).get("mid_iv",0) or 0) for o in opts_sorted if o.get("greeks")]
                avg_iv     = round(sum(ivs)/len(ivs),4) if ivs else 0
                deltas     = [float(o.get("greeks",{}).get("delta",0) or 0) for o in opts_sorted if o.get("greeks")]
                net_delta  = round(sum(deltas),2)

                try:
                    exp_d = date.fromisoformat(exp)
                    days_out = (exp_d - date.today()).days
                except:
                    days_out = "?"

                # Summary row
                sum_vals = [exp, days_out, len(strikes), len(calls), len(puts),
                            total_oi, call_oi, put_oi, avg_iv, net_delta]
                for ci, val in enumerate(sum_vals, 1):
                    cell = ws_sum.cell(sum_row, ci, val)
                    cell.font = font("E6EDF3")
                    cell.fill = fill("0D1117") if sum_row%2==0 else fill("141A22")
                ws_sum.row_dimensions[sum_row].height = 16
                sum_row += 1

                # Per-expiration sheet
                ws = wb.create_sheet(exp[:31])
                ws.sheet_view.showGridLines = False

                for row in ws.iter_rows(min_row=1, max_row=len(opts_sorted)+5, min_col=1, max_col=18):
                    for cell in row: cell.fill = fill("06080B")

                # Title
                ws.merge_cells("A1:R1")
                ws.cell(1,1, f"{sym} | {exp} | {days_out} days | Spot: ${spot:.2f} | {len(strikes)} strikes | OI: {total_oi:,}")
                ws.cell(1,1).font = Font(name="Calibri", bold=True, color="F0A500", size=11)
                ws.cell(1,1).fill = fill("0A0D12")
                ws.row_dimensions[1].height = 20

                # Column headers
                hdrs = ["TYPE","STRIKE","LAST","BID","ASK","MID","VOLUME","OI",
                        "DELTA","GAMMA","THETA","VEGA","RHO","IV (MID)","IV (BID)","IV (ASK)","INTRINSIC","TIME VAL"]
                widths = [6,10,8,8,8,8,10,10,8,8,8,8,7,10,10,10,10,10]
                for ci,(h,w) in enumerate(zip(hdrs,widths),1):
                    cell = ws.cell(2, ci, h)
                    cell.font = Font(name="Calibri", bold=True, color="6E7681", size=8)
                    cell.fill = fill("0D1117")
                    from openpyxl.utils import get_column_letter
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.row_dimensions[2].height = 16

                row_num = 3
                for o in opts_sorted:
                    try:
                        strike = float(o.get("strike",0))
                        otype  = o.get("option_type","").upper()
                        g      = o.get("greeks") or {}
                        is_call = otype == "CALL"
                        atm = abs(strike - spot) / spot < 0.02

                        if atm:
                            bg = "1A2E1A" if is_call else "2E1A1A"
                        elif is_call:
                            bg = "0A1A0A" if row_num%2==0 else "0D200D"
                        else:
                            bg = "1A0A0A" if row_num%2==0 else "200D0D"

                        txt_color = "00CC66" if is_call else "FF4455"
                        if atm: txt_color = "00FF88" if is_call else "FF6677"

                        def sf(key, dec=2):
                            v = o.get(key) if key in o else g.get(key)
                            if v is None or v == "": return ""
                            try: return round(float(v), dec)
                            except: return ""

                        last = sf("last"); bid = sf("bid"); ask = sf("ask")
                        mid  = round((float(bid or 0)+float(ask or 0))/2, 2) if bid and ask else ""

                        # Intrinsic value
                        if is_call:
                            intrinsic = max(0, spot - strike)
                        else:
                            intrinsic = max(0, strike - spot)
                        intrinsic = round(intrinsic, 2)
                        time_val  = round(float(last or 0) - intrinsic, 2) if last else ""

                        row_data = [
                            otype, strike,
                            sf("last"), sf("bid"), sf("ask"), mid,
                            sf("volume",0), sf("open_interest",0),
                            sf("delta",4) if "delta" in g else "",
                            sf("gamma",5) if "gamma" in g else "",
                            sf("theta",4) if "theta" in g else "",
                            sf("vega",4)  if "vega"  in g else "",
                            sf("rho",4)   if "rho"   in g else "",
                            sf("mid_iv",4)   if "mid_iv"   in g else "",
                            sf("bid_iv",4)   if "bid_iv"   in g else "",
                            sf("ask_iv",4)   if "ask_iv"   in g else "",
                            intrinsic, time_val,
                        ]
                        rf = fill(bg)
                        for ci, val in enumerate(row_data, 1):
                            cell = ws.cell(row_num, ci, val)
                            cell.font = Font(name="Calibri", color=txt_color, size=9)
                            cell.fill = rf
                        ws.row_dimensions[row_num].height = 14
                        row_num += 1
                    except:
                        continue

                ws.freeze_panes = "A3"

            except Exception as e:
                continue

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{sym}_OptionChain_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__=='__main__':
    app.run(debug=True,host='0.0.0.0',port=5000)
